"""
Утилиты для экспорта данных в Excel
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime


def export_table_to_excel(table_widget, file_path: str, title: str = None):
    """
    Экспорт QTableWidget в Excel файл
    
    Args:
        table_widget: QTableWidget с данными
        file_path: Путь для сохранения файла
        title: Заголовок отчета (опционально)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Данные"
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    current_row = 1
    
    # Заголовок отчета
    if title:
        ws.cell(row=current_row, column=1, value=title)
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        current_row += 1
        
        ws.cell(row=current_row, column=1, value=f"Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
        ws.cell(row=current_row, column=1).font = Font(italic=True, color="666666")
        current_row += 2
    
    # Заголовки колонок
    col_count = table_widget.columnCount()
    for col in range(col_count):
        header_item = table_widget.horizontalHeaderItem(col)
        header_text = header_item.text() if header_item else f"Колонка {col+1}"
        
        cell = ws.cell(row=current_row, column=col + 1, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    current_row += 1
    
    # Данные
    for row in range(table_widget.rowCount()):
        # Пропускаем скрытые строки
        if table_widget.isRowHidden(row):
            continue
            
        for col in range(col_count):
            item = table_widget.item(row, col)
            value = item.text() if item else ""
            
            # Пытаемся преобразовать числа
            try:
                if '.' in value:
                    value = float(value)
                elif value.isdigit() or (value.startswith('-') and value[1:].isdigit()):
                    value = int(value)
            except (ValueError, AttributeError):
                pass
            
            cell = ws.cell(row=current_row, column=col + 1, value=value)
            cell.border = thin_border
            
            # Выравнивание чисел вправо
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
        
        current_row += 1
    
    # Автоширина колонок
    for col in range(1, col_count + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        
        for row in range(1, current_row):
            cell = ws.cell(row=row, column=col)
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем файл
    wb.save(file_path)


def export_report_to_excel(headers: list, data: list, file_path: str, 
                           title: str = None, totals: dict = None):
    """
    Экспорт данных отчета в Excel
    
    Args:
        headers: Список заголовков колонок
        data: Список строк данных (каждая строка - список значений)
        file_path: Путь для сохранения
        title: Заголовок отчета
        totals: Словарь с итогами (индекс колонки -> значение)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет"
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
    totals_fill = PatternFill(start_color="FFF8F8", end_color="FFF8F8", fill_type="solid")
    totals_font = Font(bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    current_row = 1
    
    # Заголовок отчета
    if title:
        ws.cell(row=current_row, column=1, value=title)
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        ws.merge_cells(start_row=current_row, start_column=1, 
                       end_row=current_row, end_column=len(headers))
        current_row += 1
        
        ws.cell(row=current_row, column=1, 
                value=f"Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
        ws.cell(row=current_row, column=1).font = Font(italic=True, color="666666")
        current_row += 2
    
    # Заголовки колонок
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    current_row += 1
    
    # Данные
    for row_data in data:
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.border = thin_border
            
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
        
        current_row += 1
    
    # Итоговая строка
    if totals:
        for col, value in totals.items():
            cell = ws.cell(row=current_row, column=col + 1, value=value)
            cell.font = totals_font
            cell.fill = totals_fill
            cell.border = thin_border
            
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
        
        # Заполняем пустые ячейки в итоговой строке
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col)
            if cell.value is None:
                cell.fill = totals_fill
                cell.border = thin_border
    
    # Автоширина колонок
    for col in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        
        for row in range(1, current_row + 1):
            cell = ws.cell(row=row, column=col)
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(file_path)
